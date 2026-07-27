from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

INPUT = Path("content/knowledge/k.xlsx")
OUTPUT = Path("content/knowledge/特別教室使用クラス.md")

wb = load_workbook(INPUT, data_only=True)
ws = wb.active

# 結合セルを展開
merged_value = {}

for merged in ws.merged_cells.ranges:
    value = ws.cell(merged.min_row, merged.min_col).value
    for row in range(merged.min_row, merged.max_row + 1):
        for col in range(merged.min_col, merged.max_col + 1):
            merged_value[(row, col)] = value

def value(row: int, col: int):
    v = ws.cell(row, col).value
    if v is not None:
        return v
    return merged_value.get((row, col))

# 日付列を取得
dates = []

current_month = None

for col in range(4, ws.max_column + 1):
    month = value(2, col)
    if month:
        current_month = str(month).replace("月", "")

    day = value(3, col)
    if day is None:
        continue

    dates.append(
        {
            "column": col,
            "month": int(current_month),
            "day": int(day),
        }
    )

# 使用予定を抽出
records = []

current_room = None

for row in range(4, ws.max_row + 1):

    room = value(row, 1)
    if room:
        current_room = str(room)

    slot = value(row, 3)

    if slot is None:
        continue

    for d in dates:

        cls = value(row, d["column"])

        if cls in (None, ""):
            continue

        records.append(
            (
                d["month"],
                d["day"],
                str(slot),
                current_room,
                str(cls),
            )
        )

# Markdown出力
schedule = defaultdict(lambda: defaultdict(list))

for month, day, slot, room, cls in records:
    schedule[(month, day)][slot].append((room, cls))

slot_order = {
    "am1": 0,
    "am2": 1,
    "pm3": 2,
    "pm4": 3,
}

with OUTPUT.open("w", encoding="utf-8") as f:
    f.write(
"""---
title: 特別教室使用クラス
context: "最新版の第94回創作展の特別教室使用クラス一覧です。"
---

# 特別教室使用クラス

## 使用教室とその階数の対応表

| 教室 | 階数 |
| --- | --- |
| 401 | 4F |
| 第一call | 4F |
| 301 | 3F |
| 302 | 3F |
| 303 | 3F |
| 37 | 3F |
| 38 | 3F |
| 22 | 2F |
| 視聴覚 | 2F |
| 会議 | 1F |
| 多目的 | 1F |

## 項目名と時間の対応表

| 項目名 | 時間 |
| --- | --- |
| am1 | 9:00～10:30 |
| am2 | 10:30～12:00 |
| pm3 | 13:00～14:30 |
| pm4 | 14:30～16:00 |

## 使用予定

"""
    )

    for (month, day) in sorted(schedule):
        f.write(f"### {month}/{day}\n\n")

        for slot in sorted(schedule[(month, day)], key=lambda s: slot_order[s]):
            f.write(f"#### {slot}\n\n")

            for room, cls in sorted(schedule[(month, day)][slot]):
                f.write(f"- {room}: {cls}\n")

            f.write("\n")

print(f"{len(records)}件出力しました。")